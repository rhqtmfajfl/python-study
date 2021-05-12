import openpyxl as xl

exf = xl.load_workbook('c:\\dd\\itx.xlsx')
aka = exf.active
tot = 0
avg = 0
for i in aka.rows:
    index = i[0].column
    com = i[0].value
    sal = i[1].value

    tot += sal
    avg = tot / 5
    aka.cell(row = 6, column = 2).value = tot
    aka.cell(row = 7, column = 2).value = avg
    print('{} {} {} {:.2f}'.format(com, sal, tot, avg))

exf.save('c:\\dd\\outitx1.xlsx') 
exf.close()