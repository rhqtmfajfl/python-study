import openpyxl as xl

exf =  xl.load_workbook('c:\\dd\\ess.xlsx') 
aws = exf.active

for i in aws.rows:
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    mat = i[3].value

    tot = kor + eng +mat
    avg = tot /3

    aws.cell(row = index, column = 5).value = tot
    aws.cell(row = index, column = 6).value = avg

    print('{} {} {} {} {} {:.2f}'.format(name, kor, eng, mat, tot, avg))

print(dir(xl))

exf.save('c:\\dd\\outss.xlsx')    
exf.close()



